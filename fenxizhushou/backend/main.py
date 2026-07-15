import os
import re
import json
import math
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta
from dotenv import load_dotenv
import pymysql
import httpx
import pycountry
from fastapi import FastAPI, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from cachetools import TTLCache

load_dotenv()

FEED_CACHE = TTLCache(maxsize=10, ttl=7200)
SNAPSHOT_DIR = Path(__file__).resolve().parent.parent / "feed_snapshots"
SNAPSHOT_DIR.mkdir(exist_ok=True)

app = FastAPI(title="数据分析助手 API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

FRONTEND_DIR = Path(__file__).resolve().parent.parent / "frontend"

def get_conn():
    return pymysql.connect(
        host=os.environ["DB_HOST"],
        port=int(os.environ["DB_PORT"]),
        user=os.environ["DB_USER"],
        password=os.environ["DB_PASSWORD"],
        database=os.environ["DB_NAME"],
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
    )

def extract_offer_id(offer_string):
    if not offer_string:
        return ""
    match = re.match(r"^(\d+)", str(offer_string))
    return match.group(1) if match else str(offer_string)

@app.get("/api/report")
def get_report(
    date_from: str = Query(default=None, description="开始日期 YYYYMMDD"),
    date_to: str = Query(default=None, description="结束日期 YYYYMMDD"),
):
    if not date_to:
        date_to = datetime.utcnow().strftime("%Y%m%d")
    if not date_from:
        dt = datetime.strptime(date_to, "%Y%m%d") - timedelta(days=7)
        date_from = dt.strftime("%Y%m%d")

    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT date, mid, src, adgroup_id, src_offer_id, pkg_name,
                          adv_country, revenue, payout, click, conversion,
                          callback_conversion, revenue_ecpc, cvr
                   FROM offerplus_detail_report
                   WHERE date >= %s AND date <= %s
                   ORDER BY date DESC""",
                (date_from, date_to),
            )
            rows = cur.fetchall()
    finally:
        conn.close()

    result = []
    for r in rows:
        offer = r["adgroup_id"]
        result.append({
            "Date": r["date"],
            "Pub ID": r["mid"],
            "Advertiser ID": r["src"],
            "Offer": offer,
            "Adv Offer ID": r["src_offer_id"],
            "Package Name": r["pkg_name"],
            "Adv GEO": r["adv_country"],
            "Revenue": float(r["revenue"]) if r["revenue"] is not None else 0,
            "Payout": float(r["payout"]) if r["payout"] is not None else 0,
            "Click": int(r["click"]) if r["click"] is not None else 0,
            "Conversion": int(r["conversion"]) if r["conversion"] is not None else 0,
            "Callback Conversion": int(r["callback_conversion"]) if r["callback_conversion"] is not None else 0,
            "Revenue eCPC": float(r["revenue_ecpc"]) if r["revenue_ecpc"] is not None else 0,
            "CVR": float(r["cvr"]) if r["cvr"] is not None else 0,
            "OfferIdRaw": extract_offer_id(offer),
        })

    return result

@app.get("/api/dates")
def get_dates():
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT DISTINCT date FROM offerplus_detail_report ORDER BY date DESC"
            )
            rows = cur.fetchall()
    finally:
        conn.close()
    return [r["date"] for r in rows]

@app.get("/api/latest")
def get_latest():
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT MAX(date) as latest_date, MAX(created_at) as latest_created FROM offerplus_detail_report"
            )
            row = cur.fetchone()
    finally:
        conn.close()
    return {
        "latest_date": row["latest_date"],
        "latest_created": str(row["latest_created"]) if row["latest_created"] else None,
    }

@app.get("/api/pubnames")
def get_pub_names():
    return load_pub_config()

@app.get("/api/advnames")
def get_adv_names():
    mapping_str = os.environ.get("ADV_MAPPING", "")
    if mapping_str:
        try:
            return json.loads(mapping_str)
        except (json.JSONDecodeError, TypeError):
            pass
    mapping_path = Path(__file__).resolve().parent.parent / "adv_mapping.json"
    if mapping_path.exists():
        with open(mapping_path, encoding="utf-8") as f:
            return json.load(f)
    return {}

# ===== Offer 匹配 =====

def load_pub_config():
    """读取 pub_mapping 并标准化为 {name, token} 格式"""
    mapping_str = os.environ.get("PUB_MAPPING", "")
    data = None
    if mapping_str:
        try:
            data = json.loads(mapping_str)
        except (json.JSONDecodeError, TypeError):
            pass
    if data is None:
        mapping_path = Path(__file__).resolve().parent.parent / "pub_mapping.json"
        if mapping_path.exists():
            with open(mapping_path, encoding="utf-8") as f:
                data = json.load(f)
    if data is None:
        return {}
    result = {}
    for k, v in data.items():
        if isinstance(v, str):
            result[k] = {"name": v, "token": ""}
        else:
            result[k] = {"name": v.get("name", ""), "token": v.get("token", "")}
    return result

@app.get("/api/feed-snapshot")
def get_feed_snapshot(pub_id: str = Query(...)):
    path = SNAPSHOT_DIR / f"{pub_id}.json"
    if not path.exists():
        return {"exists": False}
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return {"exists": True, "timestamp": data.get("timestamp"), "count": len(data.get("offers", []))}

@app.post("/api/feed-snapshot")
async def update_feed_snapshot(pub_id: str = Query(...)):
    pub_config = load_pub_config()
    info = pub_config.get(pub_id, {})
    token = info.get("token", "") if isinstance(info, dict) else ""
    if not token:
        return {"error": "no token"}
    # 删旧快照，强制重新拉取
    snap_path = SNAPSHOT_DIR / f"{pub_id}.json"
    if snap_path.exists():
        snap_path.unlink()
    FEED_CACHE.clear()
    offers = await fetch_feed_all(pub_id, token)
    if isinstance(offers, dict) and "error" in offers:
        return offers
    path = SNAPSHOT_DIR / f"{pub_id}.json"
    data = {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"), "offers": offers}
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    return {"exists": True, "timestamp": data["timestamp"], "count": len(offers)}

async def fetch_feed_all(pub_id: str, token: str):
    """全量拉取 Feed API，优先用快照，24h 内有效"""
    cache_key = f"feed_{pub_id}"
    if cache_key in FEED_CACHE:
        return FEED_CACHE[cache_key]
    # 尝试读快照
    snap_path = SNAPSHOT_DIR / f"{pub_id}.json"
    if snap_path.exists():
        with open(snap_path, encoding="utf-8") as f:
            data = json.load(f)
        ts = datetime.strptime(data["timestamp"], "%Y-%m-%d %H:%M")
        if (datetime.now() - ts).total_seconds() < 86400:
            FEED_CACHE[cache_key] = data["offers"]
            return data["offers"]

    url = "http://doubleint.api.offerplus.net/feed/"
    page = 1
    all_offers = []

    async with httpx.AsyncClient(timeout=30) as client:
        r = await client.get(url, params={"pub_id": pub_id, "token": token, "page": 1, "per_page": 1000})
        data = r.json()
        if int(data.get("code", 0)) not in (0, 200):
            return {"error": f"Feed API error: {data.get('message', 'unknown')}"}
        total = int(data["data"]["total"])
        all_offers.extend(data["data"]["offer"])
        total_pages = math.ceil(total / 1000)
        for p in range(2, total_pages + 1):
            r = await client.get(url, params={"pub_id": pub_id, "token": token, "page": p, "per_page": 1000})
            d = r.json()
            if int(d.get("code", 0)) in (0, 200):
                all_offers.extend(d["data"]["offer"])

    FEED_CACHE[cache_key] = all_offers
    return all_offers


def build_index(offers):
    """构建 pkg_name -> {geo_code -> [offer_id, ...]}"""
    idx = {}
    for o in offers:
        pkg = (o.get("pkg_name") or "").strip()
        oid = str(o.get("offer_id", "")).strip()
        if not pkg or not oid:
            continue
        countries = (o.get("country") or "").strip().upper().split("|")
        if pkg not in idx:
            idx[pkg] = {}
        for c in countries:
            c = c.strip()
            if not c:
                continue
            if c not in idx[pkg]:
                idx[pkg][c] = []
            idx[pkg][c].append(oid)
    return idx


def normalize_geo(geo):
    """ISO alpha-3 → alpha-2 转换，如 IND→IN"""
    g = (geo or "").strip().upper()
    if len(g) == 3:
        try:
            c = pycountry.countries.get(alpha_3=g)
            if c and c.alpha_2:
                return c.alpha_2
        except Exception:
            pass
    return g

def match_offers_impl(pairs, feed_data, conn, date_from, date_to, sort_mode="ecpc", min_rev=1, min_ecpc=0.2, min_cvr=0.2):
    """三态匹配 + DB 指标查询。sort_mode: ecpc | feed"""
    if not date_to:
        date_to = datetime.utcnow().strftime("%Y%m%d")
    if not date_from:
        date_from = (datetime.strptime(date_to, "%Y%m%d") - timedelta(days=7)).strftime("%Y%m%d")
    idx = build_index(feed_data)
    results = []
    all_matched_ids = []
    for pair in pairs:
        pkg = (pair.get("pkg","") or "").strip()
        geo = normalize_geo(pair.get("geo",""))
        entry = {"pkg": pkg, "geo": geo, "status": "red", "offers": [], "other_offers": []}
        if pkg in idx:
            if geo in idx[pkg]:
                entry["status"] = "green"
                entry["offers"] = idx[pkg][geo][:3] if sort_mode == "feed" else idx[pkg][geo]
                all_matched_ids.extend(entry["offers"])
            else:
                entry["status"] = "yellow"
                for c, ids in idx[pkg].items():
                    for oid in ids[:3]:
                        entry["other_offers"].append(f"{oid}-{c}")
        results.append(entry)

    id_to_metrics = {}
    id_to_adv = {}
    cur = conn.cursor(pymysql.cursors.DictCursor)

    # 1. 查指标（按日期范围）
    if all_matched_ids:
        placeholders = ",".join(["%s"] * len(all_matched_ids))
        cur.execute(f"""SELECT adgroup_id, SUM(revenue) as rev, SUM(payout) as pay, SUM(click) as clk,
                        SUM(conversion) as conv
                        FROM offerplus_detail_report
                        WHERE adgroup_id IN ({placeholders})
                        AND date >= %s AND date <= %s
                        GROUP BY adgroup_id""",
                    (*all_matched_ids, date_from, date_to))
        for row in cur.fetchall():
            oid = str(row["adgroup_id"])
            clk = int(row["clk"] or 0)
            id_to_metrics[oid] = {
                "rev": round(float(row["rev"] or 0), 2),
                "click": clk,
                "ecpc": round((float(row["pay"] or 0) / clk) * 1000 if clk > 0 else 0, 2),
                "cvr": round((float(row["conv"] or 0) / clk) * 100 if clk > 0 else 0, 2),
            }

    # 2. 查 Adv 名（不限日期，取最先出现的 src）
    if all_matched_ids:
        ids_str = ",".join(all_matched_ids)
        cur.execute(f"SELECT adgroup_id, MIN(updated_at), src FROM offerplus_detail_report WHERE adgroup_id IN ({ids_str}) GROUP BY adgroup_id")
        for row in cur.fetchall():
            id_to_adv[str(row["adgroup_id"])] = str(row["src"] or "")

    adv_map = {}
    try:
        am_path = Path(__file__).resolve().parent.parent / "adv_mapping.json"
        if am_path.exists():
            with open(am_path, encoding="utf-8") as f:
                adv_map = json.load(f)
    except Exception:
        pass

    for entry in results:
        for i, oid in enumerate(entry["offers"]):
            m = id_to_metrics.get(oid, {})
            adv_id = id_to_adv.get(oid, "")
            adv_name = adv_map.get(adv_id, adv_id)
            entry["offers"][i] = {"id": oid, "adv": adv_name, "rev": m.get("rev", 0), "click": m.get("click", 0),
                                  "ecpc": m.get("ecpc", 0), "cvr": m.get("cvr", 0)}
        if sort_mode == "ecpc":
            entry["offers"] = sorted(entry["offers"], key=lambda o: o["ecpc"], reverse=True)[:3]
        entry["offers"] = [o for o in entry["offers"] if o["rev"] >= min_rev and o["ecpc"] >= min_ecpc and o["cvr"] >= min_cvr][:3]

    stats = {"total": len(results), "green": sum(1 for r in results if r["status"] == "green"),
             "yellow": sum(1 for r in results if r["status"] == "yellow"),
             "red": sum(1 for r in results if r["status"] == "red")}
    return results, stats


def generate_excel(results, stats, sep=",", checked_ids=None):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "匹配结果"
    checked = set(checked_ids or [])

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ["包名", "GEO", "Offer ID", "其他国家Offer ID"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row = 2
    for r in results:
        fill = green_fill if r["status"] == "green" else yellow_fill if r["status"] == "yellow" else red_fill
        if r["status"] == "green":
            ids = [o["id"] for o in r["offers"] if not checked or o["id"] in checked]
            # Skip row entirely if checked mode and no selected IDs in this row
            if checked and not ids:
                continue
            oid_str = sep.join(ids) if ids else ""
        else:
            # Skip yellow/red rows in checked mode (no offers to select)
            if checked:
                continue
            oid_str = "无匹配" if r["status"] == "red" else ""
        other = ", ".join(r["other_offers"]) if r["other_offers"] else ""
        vals = [r["pkg"], r["geo"], oid_str, other]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        row += 1

    ws.column_dimensions['A'].width = 48
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 65

    # 统计 sheet
    ws2 = wb.create_sheet("统计")
    ws2.cell(row=1, column=1, value="总计").font = Font(bold=True)
    ws2.cell(row=1, column=2, value=stats["total"])
    ws2.cell(row=2, column=1, value="🟢 绿色").font = Font(color="006100")
    ws2.cell(row=2, column=2, value=stats["green"])
    ws2.cell(row=3, column=1, value="🟡 黄色").font = Font(color="9C6500")
    ws2.cell(row=3, column=2, value=stats["yellow"])
    ws2.cell(row=4, column=1, value="🔴 红色").font = Font(color="9C0006")
    ws2.cell(row=4, column=2, value=stats["red"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.post("/api/match-offers")
@app.get("/api/match-offers")
async def match_offers_endpoint(pub_id: str = Query(...), request: Request = None):
    if request and request.method == "POST":
        body = await request.json()
        pairs = body.get("pairs", [])
        fmt = body.get("format", "")
        date_from = body.get("date_from", "")
        date_to = body.get("date_to", "")
        sep = body.get("sep", ",")
        checked_ids = body.get("checked_ids", [])
        sort_mode = body.get("sort_mode", "ecpc")
        min_rev = float(body.get("min_rev", 1))
        min_ecpc = float(body.get("min_ecpc", 0.2))
        min_cvr = float(body.get("min_cvr", 0.2))
    else:
        pairs_raw = request.query_params.get("pairs", "[]") if request else "[]"
        try:
            pairs = json.loads(pairs_raw)
        except Exception:
            pairs = []
        fmt = request.query_params.get("format", "") if request else ""
        date_from = request.query_params.get("date_from", "") if request else ""
        date_to = request.query_params.get("date_to", "") if request else ""
        sep = request.query_params.get("sep", ",") if request else ","
        checked_ids = request.query_params.get("checked_ids", "").split(",") if request else []

    if not pub_id:
        return {"error": "pub_id required"}

    pub_config = load_pub_config()
    pub_info = pub_config.get(pub_id, {})
    token = pub_info.get("token", "") if isinstance(pub_info, dict) else ""
    if not token:
        return {"error": f"pub_id {pub_id} has no token configured"}

    offers = await fetch_feed_all(pub_id, token)
    if isinstance(offers, dict) and "error" in offers:
        return offers

    conn = get_conn()
    try:
        results, stats = match_offers_impl(pairs, offers, conn, date_from, date_to, sort_mode, min_rev, min_ecpc, min_cvr)
    finally:
        conn.close()

    if fmt == "xlsx":
        buf = generate_excel(results, stats, sep, checked_ids)
        filename = f"offer_match_{pub_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename={filename}"})

    return {"pub_id": pub_id, "stats": stats, "matches": results}

if FRONTEND_DIR.exists():
    app.mount("/", StaticFiles(directory=str(FRONTEND_DIR), html=True))
